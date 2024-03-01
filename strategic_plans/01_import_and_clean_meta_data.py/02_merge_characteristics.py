# %%
import pandas as pd
import re
from strategic_plans.library import start
import numpy as np
from openpyxl import load_workbook

# %%
df = pd.read_csv(start.MAIN_DIR + "data/clean/plans_meta_data.csv")
df = df[
    [
        "leaid",
        "plan_identified",
        "dedoose_name",
        "need_to_code",
        "plan_complete",
        "priority",
    ]
]
covars = pd.read_csv(start.DATA_DIR + "clean/seda_ccd_covariates_2018.csv")

# %%
df = df.merge(covars, on="leaid", how="left", indicator="_merge_covars")
df._merge_covars.value_counts()
df["leaid"] = df.leaid.astype(int)
# %%

seda = pd.read_csv(start.DATA_DIR + "clean/2018_seda_outcomes.csv")

df = df.merge(seda, on="leaid", how="left", indicator="_merge_seda")
df._merge_seda.value_counts()
# %%


# %%
characteristics = [
    "perasn",
    "perblk",
    "perhsp",
    "perfrl",
    "perell",
    "lninc50all",
    "test_rla_all_mean",
    "test_math_all_mean",
]

df.to_csv(start.MAIN_DIR + "data/clean/plans_meta_data_full.csv")

# %%
# locations
file = start.MAIN_DIR + "results/Sample Characteristics.xlsx"
wb = load_workbook(file)
ws = wb["Sheet1"]

# %%
plan_df = df[df.plan_identified == 1]
qual_df = df[df.priority == 1]
missing_df = df[df.plan_identified != 1]

# locales
col = 2
for sample_df in [df, plan_df, qual_df, missing_df]:
    row = 2
    for locale in ["urban", "suburb", "town", "rural"]:
        ws.cell(row=row, column=col).value = sample_df[locale].mean().round(2)
        row = row + 1

    for count_var in ["operational_schools", "totenrl"]:
        ws.cell(row=row, column=col).value = sample_df[count_var].mean().round(0)
        row = row + 1

    for char in characteristics:
        ws.cell(row=row, column=col).value = sample_df[char].mean().round(2)
        row = row + 1

    ws.cell(row=row, column=col).value = len(sample_df)
    col = col + 1


wb.save(file)

# %%
